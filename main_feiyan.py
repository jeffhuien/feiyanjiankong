# -*- coding: utf-8 -*-
# Author: GAO--HUI
# Date: 2022-01-23 13:08:32
# LastEditors: GAO--HUI
# LastEditTime: 2022-02-16 11:06:32
# FilePath: \疫情\main_feiyan.py


import re
import time
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from selenium import webdriver


def download_html():
    # 1.访问百度
    drive = webdriver.Edge(executable_path=("./edgedriver_win64/msedgedriver.exe"))
    drive.get("https://news.qq.com/zt2020/page/feiyan.htm#/")
    # 3.休眠2s,获取服务器的响应内容
    time.sleep(1)
    # 4.获取页面源码数据
    text = drive.page_source
    # 将数据写入文件
    with open("html/肺炎_{}.html".format(time.strftime("%Y_%m_%d")), "w", encoding="u8") as f:
        f.write(text)
    # 关闭浏览器窗口
    drive.close()
    return text


def write_xls(data: list):
    doc = openpyxl.load_workbook("肺炎疫情历史情况.xlsx")  # 加载工作表
    # doc.active = doc.get_sheet_names()[0]
    if time.strftime("%Y_%m_%d") in doc.get_sheet_names():
        doc.remove(doc[time.strftime("%Y_%m_%d")])
    xls = doc.create_sheet(time.strftime("%Y_%m_%d"), 0)  # 创建当天的工作表
    doc.active = xls

    # 写入数据
    for i in range(len(data)):
        xls.append(data[i])
        # 找到河南数据行，设置颜色为红色
        if "河南" in data[i]:
            for j in list(xls.rows)[-1]:
                print(j.value)
                j.font = Font(color="ff0000", bold=True)

    xls["A1"].font = Font(color="ff0000", bold=True)

    for i in range(1, xls.max_column + 1):
        print(get_column_letter(i))
        if get_column_letter(i) not in list("AD"):
            xls.column_dimensions[get_column_letter(i)].width = 8 + 0.62
        else:
            xls.column_dimensions[get_column_letter(i)].width = 12 + 0.62
        print(xls.column_dimensions[get_column_letter(i)].width)
        xls.cell(row=12, column=i).alignment = Alignment(horizontal="center")
        xls.cell(row=12, column=i).font = Font(color="ff0000", bold=True)

    # 设置标签页颜色  会报错但是不影响使用
    for i in doc.get_sheet_names()[0::2]:
        doc[i].sheet_properties.tabColor = "ff72BA"
    for i in doc.get_sheet_names()[1::2]:
        doc[i].sheet_properties.tabColor = "00B0F0"

    doc.save("肺炎疫情历史情况.xlsx")
    doc.close()


def main():
    # htmlpage = download_html()
    download_html()
    with open("html/肺炎_{}.html".format(time.strftime("%Y_%m_%d")), "r", encoding="u8") as f:
        # with open("html/肺炎_{}.html".format("2022_01_29"), "r", encoding="u8") as f:
        htmlpage = f.read()
    dailybox = []
    time1 = re.findall(r'<div data-v-7fcb7d83="" class="timeNum">.*?<p data-v-7fcb7d83="" class="d"> 统计截至 <span data-v-7fcb7d83="">(.*?)</span><em', htmlpage, re.S)[-1]
    databox = [["统计截至：" + time1], [""], ["国内疫情总览"]]

    dailytext = re.findall(r'<div data-v-7fcb7d83="" class="timeNum">(.*?亡</div>)', htmlpage, re.S)[-1]

    boxnum = re.findall(r'<div data-v-7fcb7d83="" class="number">(.*?)</div>', dailytext, re.S)
    for i in range(len(boxnum)):
        boxnum[i] = int(boxnum[i])

    boxtxt = re.findall(
        r'<div data-v-7fcb7d83="" class="add"> (较上日)<span data-v-7fcb7d83="">(.*?)</span></div>',
        dailytext,
        re.S,
    )
    for i in range(len(boxtxt)):
        boxtxt[i] = "".join(boxtxt[i])

    boxtitile = re.findall(
        r'<div data-v-7fcb7d83="" class="text"><span data-v-7fcb7d83="">(.*?)</span></div>',
        dailytext,
        re.S,
    )
    dailytext = dailytext[dailytext.find('<div data-v-7fcb7d83="" class="icbar confirm">') : :]
    boxtitile.extend(
        re.findall(
            r'<div data-v-7fcb7d83="" class="text">(.*?)</div>',
            dailytext,
            re.S,
        )
    )

    for i in range(6):
        dailybox.append([boxtitile[i], boxnum[i], boxtxt[i]])

    databox.extend(dailybox)

    databox.append(
        [
            "",
        ]
    )

    databox.append(
        [
            "中国疫情（包括港澳台）",
        ]
    )
    databox.append(
        [
            "地区",
            "现有",
            "累计",
            "",
            "治愈",
            "死亡",
        ]
    )

    # 获取数据块
    data = []
    for i in range(len(re.findall(r'<tr data-v-4eb96304="" class="areaBox">', htmlpage, re.S))):
        data = re.findall(r'<tbody data-v-4eb96304="" class="">(.*?)<!---->', htmlpage, re.S)

    for i in data:
        areadata = []
        area = re.findall(r'<p data-v-4eb96304=""><span data-v-4eb96304="">(.*?)</span></p>', i)
        areadata.append(area[-1])
        existing = re.findall(r'<p data-v-4eb96304="" class="bold">(.*?)</p>', i)
        for j in range(len(existing)):
            existing[j] = int(existing[j])
        existing.insert(2, re.findall(r'<p data-v-4eb96304=""> (.*?) </p>', i)[-1])
        areadata.extend(existing)
        databox.append(areadata)
        print(databox[-1])
    # region
    # endregion

    return databox


if __name__ == "__main__":
    write_xls(main())
    # a = input()
    # main()
